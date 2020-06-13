from openpyxl import load_workbook

DEBUG = False
RUN_DEBUG = False
if __name__ == '__main__':
    RUN_DEBUG = True

DEFAULT_SETTINGS = {
  "key_column": "C",
  "result": {
    "name": "F",
    "abv": "G",
    "ibu": "H",
    "style": "I",
    "brewery": "J",
    "country": "K"
  }
}


def read_xlsx(filename, column_name, settings):
    wb = load_workbook(filename)
    sheet = wb.sheetnames[0]
    table = wb[sheet]
    names = table[column_name]
    result = []
    for name_cell in names:
        name = name_cell.value
        row = name_cell.row
        if DEBUG:
            print(name)
        if name is None:
            continue
        have_all = True
        for column in settings.values():
            if table[column+str(row)].value is None:
                have_all = False
        if not have_all:
            result.append(name)
    return result

if RUN_DEBUG:
    column_name = 'C'
    filename = '/home/pronin-ao/Downloads/Ostatki_DrimkasUchet_1.xlsx'
    res = read_xlsx(filename, column_name, DEFAULT_SETTINGS["result"])
    print(res)
