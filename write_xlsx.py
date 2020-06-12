from openpyxl import load_workbook

DEBUG = True
RUN_DEBUG = False
if __name__ == '__main__':
    RUN_DEBUG = True


def write_xlsx(filename, key_column, columns, data):
    wb = load_workbook(filename)
    sheet = wb.sheetnames[0]
    table = wb[sheet]
    names = table[key_column]
    for row_key in names:
        name = row_key.value
        if name is None:
            continue
        row_num = row_key.row
        if DEBUG:
            print('process {} row with key {}'.format(row_num, name))

        if name not in data:
            if DEBUG:
                print('have not result for ', name)
            continue

        for param, letter in columns.items():
            if param not in data[name]:
                print('{} has not param {}'.format(name, param))
                continue
            table[letter+'{}'.format(row_num)] = data[name][param]
    wb.save(filename)
    print(filename, ' saved')


if RUN_DEBUG:
    columns = {
        'name': 'F',
        'style': 'H',
        'abv': 'G',
        'brewery': 'I',
        'country': 'J',
    }

    data = {
        'Altmuller Hell': {
            'name': 'HELL',
            'style': 'ALT',
            'abv': 100500,
            'brewery': 'AAA',
            'country': 'IDONNO',
        },
        '3000 IBU': {
            'name': '3000',
            'style': 'ochen gorky',
            'abv': 200300,
            'brewery': 'III',
            'country': 'US',
        },
        'ADHA 483': {
            'name': '483',
            'style': 'adha',
            'abv': 400400,
            'brewery': 'OOO',
            'country': 'SU',
        },
    }
    filename = 'Ostatki_DrimkasUchet_1.xlsx'
    key_column = 'C'
    write_xlsx(filename, key_column, columns, data)
